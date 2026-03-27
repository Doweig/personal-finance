"""One-time importer for the historical Excel file (hma-22-26.xlsx).

Reads all sheets and populates DuckDB tables: restaurants, monthly_pl,
dividends, investments, ownership.
"""

import datetime
from pathlib import Path

import openpyxl

from ingestion.db import (
    create_schema,
    insert_dividend,
    insert_investment,
    insert_ownership,
    insert_restaurant,
    upsert_monthly_pl,
)

# ---------------------------------------------------------------------------
# Restaurant metadata
# ---------------------------------------------------------------------------

RESTAURANTS = [
    {"id": "mozza-emq", "name": "Mozza EmQuartier", "restaurant_code": "17-Mozza EMQ"},
    {"id": "cocotte-39", "name": "Cocotte 39", "restaurant_code": "15-Cocotte"},
    {"id": "mozza-prg", "name": "Mozza Paragon", "restaurant_code": "20-Mozza Paragon"},
    {"id": "mozza-icsm", "name": "Mozza Icon Siam", "restaurant_code": "25-Mozza ICON"},
    {"id": "mozza-cp", "name": "Mozza Central Park", "restaurant_code": "26-Mozza Central Park"},
]

# Maps sheet-name prefix to restaurant id
SHEET_PREFIX_TO_ID = {
    "Mozza EMQ": "mozza-emq",
    "Cocotte 39": "cocotte-39",
    "Cocotte": "cocotte-39",
    "Mozza PRG": "mozza-prg",
    "Mozza ICSM": "mozza-icsm",
    "Mozza CP": "mozza-cp",
}

# Known investment amounts (THB) and ownership percentages from contracts
INVESTMENTS = {
    "mozza-emq": {"amount_thb": 12_105_000, "ownership_pct": 10.6891,
                  "date": datetime.date(2022, 9, 8)},
    "cocotte-39": {"amount_thb": 25_325_000, "ownership_pct": 21.875,
                   "date": datetime.date(2022, 10, 2)},
    "mozza-prg": {"amount_thb": 1_050_000, "ownership_pct": 1.0,
                  "date": datetime.date(2024, 6, 14)},
    "mozza-icsm": {"amount_thb": 9_000_000, "ownership_pct": 10.0,
                   "date": datetime.date(2024, 10, 18)},
    "mozza-cp": {"amount_thb": 8_000_000, "ownership_pct": 10.0,
                 "date": datetime.date(2025, 9, 26)},
}

# P&L sheet name -> (restaurant_id, year)
PL_SHEETS = {
    "Mozza EMQ 22": ("mozza-emq", 2022),
    "Mozza EMQ 23": ("mozza-emq", 2023),
    "Mozza EMQ 24": ("mozza-emq", 2024),
    "Cocotte 22": ("cocotte-39", 2022),
    "Cocotte 23": ("cocotte-39", 2023),
    "Cocotte 24": ("cocotte-39", 2024),
}

# Month name -> month number
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4,
    "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
}


def _to_date(val):
    """Convert a cell value to datetime.date, or None."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    return None


def _parse_month_name(val, year):
    """Parse a month name string or datetime into a (year, month) date."""
    if isinstance(val, datetime.datetime):
        return datetime.date(val.year, val.month, 1)
    if isinstance(val, str):
        month_num = MONTH_MAP.get(val.strip().lower())
        if month_num:
            return datetime.date(year, month_num, 1)
    return None


def _is_numeric(val):
    """Check if a value is numeric (int or float)."""
    return isinstance(val, (int, float)) and not isinstance(val, bool)


def _num(val):
    """Return numeric value or None."""
    if _is_numeric(val):
        return float(val)
    return None


# ---------------------------------------------------------------------------
# Valuation sheet parser (monthly revenue + net income)
# ---------------------------------------------------------------------------

def _import_valuation_sheet(conn, ws, restaurant_id):
    """Import monthly revenue and net income from a valuation sheet.

    These sheets have: col A = date, col B = revenue, col C = net income.
    Data starts at row 2 and ends when we hit a non-date row.
    """
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        dt = _to_date(date_val)
        if dt is None:
            # Could be a summary label or empty row - skip
            if isinstance(date_val, str):
                break  # Hit summary block
            continue

        revenue = _num(ws.cell(row, 2).value)
        net_income = _num(ws.cell(row, 3).value)

        if revenue is None and net_income is None:
            continue

        # Use first-of-month for the P&L month key
        month_date = datetime.date(dt.year, dt.month, 1)

        upsert_monthly_pl(conn, {
            "restaurant_id": restaurant_id,
            "month": month_date,
            "revenue": revenue,
            "gop_net": net_income,
        })


# ---------------------------------------------------------------------------
# Dividend sheet parser
# ---------------------------------------------------------------------------

def _import_dividend_sheet(conn, ws, restaurant_id):
    """Import dividends from a dividend sheet.

    Col A = date, col B = total THB, col E = ownership %, col F = comment.
    Row with negative col B = initial investment (skip as dividend).
    """
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, 1).value
        dt = _to_date(date_val)
        if dt is None:
            if isinstance(date_val, str):
                break  # Hit summary block like "Total"
            continue

        total_thb = _num(ws.cell(row, 2).value)
        if total_thb is None:
            continue

        # Skip initial investment rows (negative amounts)
        if total_thb < 0:
            continue

        ownership_pct = _num(ws.cell(row, 5).value)
        comment = ws.cell(row, 6).value
        if isinstance(comment, str):
            comment = comment.strip() or None
        else:
            comment = None

        # Calculate my_share_thb: total_thb is already the investor's share
        my_share_thb = total_thb

        insert_dividend(conn, {
            "restaurant_id": restaurant_id,
            "date": dt,
            "total_thb": total_thb,
            "my_share_thb": my_share_thb,
            "comment": comment,
        })


# ---------------------------------------------------------------------------
# P&L detail sheet parser
# ---------------------------------------------------------------------------

def _find_month_columns(ws):
    """Find columns that contain month data in row 1.

    Returns list of (col_index, month_date_or_name) for each month block.
    The data value column is col_index + 1.
    """
    months = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is not None:
            months.append((c, val))
    return months


def _find_label_offset(ws, month_cols):
    """Determine the offset between the month column and label/value columns.

    In most sheets, the label is at the month column and value at month_col+1.
    In Cocotte 23, labels are at the month column and values at month_col+1.
    We detect this by checking row 4 (TOTAL REVENUE).
    """
    if not month_cols:
        return 0, 1  # label_offset, value_offset

    first_col = month_cols[0][0]
    # Check if the label "TOTAL REVENUE" is at first_col (offset 0) or first_col itself
    cell_at_col = ws.cell(4, first_col).value
    if isinstance(cell_at_col, str) and "TOTAL REVENUE" in cell_at_col.upper():
        return 0, 1  # label at col, value at col+1

    # Check col+1
    cell_at_col1 = ws.cell(4, first_col + 1).value if first_col + 1 <= ws.max_column else None
    if isinstance(cell_at_col1, str) and "TOTAL REVENUE" in cell_at_col1.upper():
        # Cocotte 23 style: month name at col, but label at col, value at col+1
        return 0, 1

    return 0, 1


def _import_pl_sheet(conn, ws, restaurant_id, year):
    """Import P&L data from a detail sheet.

    Structure: months as column groups (4 cols each: label, value, ratio, spacer).
    Row 1 has month names, rows 4-19 have the P&L data.
    """
    month_cols = _find_month_columns(ws)
    if not month_cols:
        return

    for col, month_val in month_cols:
        month_date = _parse_month_name(month_val, year)
        if month_date is None:
            continue

        # Determine where label and value are relative to col
        # Check if label is at col or col has something else
        label_cell = ws.cell(4, col).value
        if isinstance(label_cell, str) and "TOTAL REVENUE" in label_cell.upper():
            # Standard layout: label at col, value at col+1
            val_col = col + 1
        else:
            # Try: the month col itself might have the value in some layouts
            # Actually look for "TOTAL REVENUE" in nearby cells
            found = False
            for offset in range(0, 3):
                test_cell = ws.cell(4, col + offset).value
                if isinstance(test_cell, str) and "TOTAL REVENUE" in test_cell.upper():
                    val_col = col + offset + 1
                    found = True
                    break
            if not found:
                continue

        # Read P&L line items
        revenue = _num(ws.cell(4, val_col).value)
        revenue_n1 = _num(ws.cell(5, val_col).value)
        rebate = _num(ws.cell(7, val_col).value)
        food_cost = _num(ws.cell(8, val_col).value)
        beverage_cost = _num(ws.cell(9, val_col).value)
        total_fb_cost = _num(ws.cell(10, val_col).value)
        total_other_exp = _num(ws.cell(12, val_col).value)
        total_monthly_exp = _num(ws.cell(14, val_col).value)
        gop_before_fee = _num(ws.cell(15, val_col).value)
        other_special_fee = _num(ws.cell(16, val_col).value)
        monthly_provision = _num(ws.cell(17, val_col).value)
        gop_net = _num(ws.cell(19, val_col).value)

        upsert_monthly_pl(conn, {
            "restaurant_id": restaurant_id,
            "month": month_date,
            "revenue": revenue,
            "revenue_n1": revenue_n1,
            "food_cost": food_cost,
            "beverage_cost": beverage_cost,
            "total_fb_cost": total_fb_cost,
            "total_other_expenses": total_other_exp,
            "total_monthly_exp": total_monthly_exp,
            "gop_before_fee": gop_before_fee,
            "other_special_fee": other_special_fee or 0,
            "monthly_provision": monthly_provision or 0,
            "gop_net": gop_net,
            "rebate": rebate or 0,
        })


# ---------------------------------------------------------------------------
# Main import function
# ---------------------------------------------------------------------------

def import_xlsx(conn, filepath):
    """Read the Excel workbook and populate all DuckDB tables.

    Args:
        conn: DuckDB connection (schema must already exist).
        filepath: Path to hma-22-26.xlsx.
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)

    # 1. Insert restaurants
    for r in RESTAURANTS:
        insert_restaurant(conn, r)

    # 2. Insert investments and ownership from known contract data
    for rid, info in INVESTMENTS.items():
        insert_investment(conn, {
            "restaurant_id": rid,
            "date": info["date"],
            "amount_thb": info["amount_thb"],
        })
        insert_ownership(conn, {
            "restaurant_id": rid,
            "effective_date": info["date"],
            "ownership_pct": info["ownership_pct"],
        })

    # 3. Import valuation sheets (revenue + net income per month)
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith(" val."):
            prefix = sheet_name.replace(" val.", "")
            rid = SHEET_PREFIX_TO_ID.get(prefix)
            if rid:
                _import_valuation_sheet(conn, wb[sheet_name], rid)

    # 4. Import dividend sheets
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith(" div."):
            prefix = sheet_name.replace(" div.", "")
            rid = SHEET_PREFIX_TO_ID.get(prefix)
            if rid:
                _import_dividend_sheet(conn, wb[sheet_name], rid)

    # 5. Import P&L detail sheets (overwrite/enrich valuation data)
    for sheet_name, (rid, year) in PL_SHEETS.items():
        if sheet_name in wb.sheetnames:
            _import_pl_sheet(conn, wb[sheet_name], rid, year)

    wb.close()
